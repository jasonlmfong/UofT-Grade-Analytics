from openpyxl import load_workbook
from grade_to_gpa import find_gpa
from Predictor import predict

'''print(sheet['A18'].value) prints value in box A18'''
'''print(sheet.max_row) prints the max number of rows'''

def weighted_average_grade(workbook):
	"""outputs weighted average of grades"""
	total_weight = 0
	total_grade = 0

	for i in range(2, workbook.max_row+1):
		if workbook[f"{'D'}{i}"].value is False:
			if workbook[f"{'C'}{i}"].value == "IPR":
				total_grade += workbook[f"{'B'}{i}"].value * predict(workbook[f"{'A'}{i}"].value)
			else:
				total_grade += workbook[f"{'B'}{i}"].value * workbook[f"{'C'}{i}"].value 
			total_weight += workbook[f"{'B'}{i}"].value

	print("your weighted average grade is")
	print(total_grade/total_weight)

def unweighted_average_grade(workbook):
	"""outputs unweighted average of grades"""
	count = 0
	total_grade = 0

	for i in range(2, workbook.max_row+1):
		if workbook[f"{'D'}{i}"].value is False:
			if workbook[f"{'C'}{i}"].value == "IPR":
				total_grade += predict(workbook[f"{'A'}{i}"].value)
			else:
				total_grade += workbook[f"{'C'}{i}"].value
			count += 1

	print("your unweighted average grade is")
	print(total_grade/count)

def weighted_average_gpa(workbook):
	"""outputs weighted average of grades"""
	total_weight = 0
	total_gpa = 0

	for i in range(2, workbook.max_row+1):
		if workbook[f"{'D'}{i}"].value is False:
			if workbook[f"{'C'}{i}"].value == "IPR":
				total_gpa += workbook[f"{'B'}{i}"].value * find_gpa(predict(workbook[f"{'A'}{i}"].value))
			else:
				total_gpa += workbook[f"{'B'}{i}"].value * find_gpa(workbook[f"{'C'}{i}"].value)
			total_weight += workbook[f"{'B'}{i}"].value

	print("your weighted average gpa is")
	print(total_gpa/total_weight)

def unweighted_average_gpa(workbook):
	"""outputs unweighted average of grades"""
	count = 0
	total_gpa = 0

	for i in range(2, workbook.max_row+1):
		if workbook[f"{'D'}{i}"].value is False:
			if workbook[f"{'C'}{i}"].value == "IPR":
				total_gpa += find_gpa(predict(workbook[f"{'A'}{i}"].value))
			else:
				total_gpa += find_gpa(workbook[f"{'C'}{i}"].value)
			count += 1

	print("your unweighted average gpa is")
	print(total_gpa/count)

if __name__ == "__main__":
	wb = load_workbook(filename = 'Transcript.xlsx')
	sheet = wb['Sheet1']
	weighted_average_grade(sheet)
	unweighted_average_grade(sheet)
	weighted_average_gpa(sheet)
	unweighted_average_gpa(sheet)