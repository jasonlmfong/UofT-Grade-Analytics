import numpy as np
from openpyxl import load_workbook
from grade_to_gpa import find_gpa

#reads the workbook and parse out the information
def parse_syllabus(workbook):
	"""parses a syllabus and current grades"""
	sum_weight = []
	counter_weight = 0
	weights = []
	grades = []

	for i in range(2, workbook.max_row+1):
		counter_weight += workbook[f"{'A'}{i}"].value
		weights.append(workbook[f"{'A'}{i}"].value)
		sum_weight.append(counter_weight)
		grades.append(workbook[f"{'B'}{i}"].value)

	return (grades,weights,sum_weight)

def calc_curr(lst1,lst2):
	"""calculates current points obtained"""
	sum = 0

	for i in range(0,len(lst1)):
		sum += lst1[i]*lst2[i]/100
	return sum

def predict(str):
	"""use linear regression to predict grade outcome"""
	wb = load_workbook(filename = f"{str}{'.xlsx'}").active

	a,b,c = parse_syllabus(wb)
	x=np.array(c)
	y = np.array(a)
	slope, intercept = np.polyfit(x, y, 1)
	remain = 100 - c[-1] 
	pred = slope*100+intercept
	sum = calc_curr(b,a) + remain*pred/100
	return(sum)
